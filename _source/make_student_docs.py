# -*- coding: utf-8 -*-
"""Build the student hand-in set.

    python3 make_student_docs.py

Ten Word documents and three spreadsheets, into ./student-docs/.

Word and Excel because these are the ones students TYPE INTO. Upload a
.docx to Drive and Google converts it to a Doc; upload an .xlsx and it
becomes a Sheet. Everything else the shop hands out is a PDF, because
nobody types into it.
"""

import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

import student_docs as SD
import generation as G

def _unent(v):
    """The data file is shared with the website builders, which want HTML
    entities. Word wants the characters."""
    if isinstance(v, str):
        for a, b in (('&amp;', '&'), ('&rsquo;', '\u2019'),
                     ('&mdash;', '\u2014'), ('&ldquo;', '\u201c'),
                     ('&rdquo;', '\u201d'), ('&hellip;', '\u2026')):
            v = v.replace(a, b)
        return v
    if isinstance(v, (list, tuple)):
        return type(v)(_unent(x) for x in v)
    if isinstance(v, dict):
        return {k: _unent(x) for k, x in v.items()}
    return v

import student_doc_data as DATA

HERE = os.path.dirname(os.path.abspath(__file__))
OUT = os.path.join(HERE, 'student-docs')

PURPLE = '6B4785'
SOFT = 'F1EBF7'
RULE = 'C3D0DA'
INK = '141C26'
INK3 = '6D7C8A'

THIN = Side(style='thin', color=RULE)
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _sheet_head(ws, title, subtitle, ncols):
    ws['A1'] = 'BHR ENGINEERING TECHNOLOGY   \u00b7   ' + G.STAMP
    ws['A1'].font = Font(name='Calibri', size=8.5, bold=True, color=PURPLE)
    ws['A2'] = title
    ws['A2'].font = Font(name='Calibri', size=16, bold=True, color=INK)
    ws['A3'] = subtitle
    ws['A3'].font = Font(name='Calibri', size=9.5, italic=True, color='3F4E5D')
    for r in (1, 2, 3):
        ws.merge_cells(start_row=r, start_column=1, end_row=r,
                       end_column=ncols)
    ws.row_dimensions[2].height = 22
    ws.freeze_panes = 'A1'


def _header_row(ws, row, headers, widths):
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = Font(name='Calibri', size=9, bold=True, color=PURPLE)
        c.fill = PatternFill('solid', fgColor=SOFT)
        c.border = BOX
        c.alignment = Alignment(vertical='center', wrap_text=True)
        ws.column_dimensions[get_column_letter(i)].width = widths[i - 1]
    ws.row_dimensions[row].height = 26


def _blank_rows(ws, first, last, ncols, height=20):
    for r in range(first, last + 1):
        ws.row_dimensions[r].height = height
        for i in range(1, ncols + 1):
            ws.cell(row=r, column=i).border = BOX


def _identity_rows(ws, row, fields):
    for i, label in enumerate(fields):
        c = ws.cell(row=row + i, column=1, value=label)
        c.font = Font(name='Calibri', size=9, bold=True, color=PURPLE)
        c.fill = PatternFill('solid', fgColor=SOFT)
        c.border = BOX
        v = ws.cell(row=row + i, column=2)
        v.border = BOX
    return row + len(fields) + 1


# ------------------------------------------------------------- research log
def research_log(path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Research Log'
    heads = ['Date', 'Source type', 'Why this source',
             'Source (author, title, year)', 'Link', 'What you took from it']
    widths = [11, 15, 20, 38, 34, 44]
    _sheet_head(ws, 'Research Log',
                'One new source a day is the target. In brainstorming and '
                'research it should be more than that.', len(heads))
    _identity_rows(ws, 5, ['Name', 'Project'])
    ws.column_dimensions['A'].width = widths[0]
    _header_row(ws, 9, heads, widths)
    _blank_rows(ws, 10, 90, len(heads))

    lists = wb.create_sheet('Lists')
    lists['A1'] = 'Source type'
    lists['B1'] = 'Why this source'
    for c in ('A1', 'B1'):
        lists[c].font = Font(bold=True, color=PURPLE)
    types = ['Article', 'Book', 'Web video', 'TV or film', 'Podcast',
             'Datasheet or manual', 'Standard or code', 'Patent',
             'Person I spoke to', 'Misc.']
    reasons = ['Technical — skill', 'Educational — knowledge',
               'Informational', 'Inspirational', 'Motivational']
    for i, v in enumerate(types, start=2):
        lists.cell(row=i, column=1, value=v)
    for i, v in enumerate(reasons, start=2):
        lists.cell(row=i, column=2, value=v)
    lists.column_dimensions['A'].width = 24
    lists.column_dimensions['B'].width = 26

    dv1 = DataValidation(type='list',
                         formula1='=Lists!$A$2:$A$%d' % (len(types) + 1),
                         allow_blank=True)
    dv2 = DataValidation(type='list',
                         formula1='=Lists!$B$2:$B$%d' % (len(reasons) + 1),
                         allow_blank=True)
    ws.add_data_validation(dv1)
    ws.add_data_validation(dv2)
    dv1.add('B10:B90')
    dv2.add('C10:C90')
    wb.save(path)


# --------------------------------------------------------- order request form
def order_request(path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Order Request'
    heads = ['Qty', 'Item description', 'Unit cost', 'Total cost',
             'Why you need it', 'Vendor link', 'Alternative link']
    widths = [6, 40, 11, 11, 40, 34, 34]
    _sheet_head(ws, 'Order Request Form',
                'One row per item. The "why you need it" column is the one '
                'that gets the order approved.', len(heads))
    nxt = _identity_rows(ws, 5, ['Name', 'Project', 'Date submitted'])
    ws.column_dimensions['A'].width = widths[0]
    ws.column_dimensions['B'].width = widths[1]
    hrow = 10
    _header_row(ws, hrow, heads, widths)
    _blank_rows(ws, hrow + 1, hrow + 25, len(heads))
    for r in range(hrow + 1, hrow + 26):
        ws.cell(row=r, column=4).value = '=IF(A%d="","",A%d*C%d)' % (r, r, r)
        ws.cell(row=r, column=3).number_format = '$#,##0.00'
        ws.cell(row=r, column=4).number_format = '$#,##0.00'
    tot = hrow + 27
    c = ws.cell(row=tot, column=3, value='Total')
    c.font = Font(name='Calibri', size=10, bold=True, color=PURPLE)
    c.alignment = Alignment(horizontal='right')
    t = ws.cell(row=tot, column=4,
                value='=SUM(D%d:D%d)' % (hrow + 1, hrow + 25))
    t.font = Font(name='Calibri', size=11, bold=True, color=INK)
    t.number_format = '$#,##0.00'
    t.border = BOX
    ws.cell(row=tot + 2, column=1,
            value='Approved by: ________________________     '
                  'Date: ____________     PO / account: ____________'
            ).font = Font(name='Calibri', size=9, color=INK3)
    wb.save(path)


# ---------------------------------------------------------------- gantt chart
def gantt(path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Gantt'
    heads = ['WBS', 'Task title', 'Work type', 'Status', 'Start', 'Due',
             'Days']
    widths = [7, 34, 15, 14, 11, 11, 7]
    WEEKS = 10
    ncols = len(heads) + WEEKS * 5
    _sheet_head(ws, 'Project Gantt Chart',
                'Break the project into tasks, then put them on a calendar. '
                'The point is to find out where two tasks collide.', 12)

    nxt = _identity_rows(ws, 5, ['Project title', 'Project manager'])

    hrow = 9
    # week banner across the day columns
    for w in range(WEEKS):
        first = len(heads) + 1 + w * 5
        ws.merge_cells(start_row=hrow - 1, start_column=first,
                       end_row=hrow - 1, end_column=first + 4)
        c = ws.cell(row=hrow - 1, column=first, value='Week %d' % (w + 1))
        c.font = Font(name='Calibri', size=9, bold=True, color=PURPLE)
        c.alignment = Alignment(horizontal='center')
        c.fill = PatternFill('solid', fgColor=SOFT)

    _header_row(ws, hrow, heads, widths)
    for w in range(WEEKS):
        for i, d in enumerate(['M', 'T', 'W', 'R', 'F']):
            col = len(heads) + 1 + w * 5 + i
            c = ws.cell(row=hrow, column=col, value=d)
            c.font = Font(name='Calibri', size=8, bold=True, color=PURPLE)
            c.fill = PatternFill('solid', fgColor=SOFT)
            c.border = BOX
            c.alignment = Alignment(horizontal='center')
            ws.column_dimensions[get_column_letter(col)].width = 3.2

    seed = [
        ('1', 'Initiating', '', '', '', '', ''),
        ('1.1', 'Define the problem', 'Research', 'Not started', '', '', ''),
        ('1.2', 'Create design requirements', 'Written', 'Not started',
         '', '', ''),
        ('2', 'Planning', '', '', '', '', ''),
        ('2.1', 'Scope and goal setting', 'Brainstorming', 'Not started',
         '', '', ''),
        ('3', 'Design', '', '', '', '', ''),
        ('3.1', 'Conceptual design', 'Drawing', 'Not started', '', '', ''),
        ('3.2', 'Detailed design', 'CAD', 'Not started', '', '', ''),
        ('4', 'Development', '', '', '', '', ''),
        ('4.1', 'Build the prototype', 'Construction', 'Not started',
         '', '', ''),
        ('4.2', 'Test and evaluate', 'Model', 'Not started', '', '', ''),
        ('4.3', 'Optimise and redesign', 'CAD', 'Not started', '', '', ''),
        ('5', 'Closing', '', '', '', '', ''),
        ('5.1', 'Presentation', 'Written', 'Not started', '', '', ''),
    ]
    r = hrow + 1
    for row in seed:
        for i, v in enumerate(row, start=1):
            c = ws.cell(row=r, column=i, value=v or None)
            c.border = BOX
            if not row[2] and i <= 2:          # a phase heading row
                c.font = Font(name='Calibri', size=10, bold=True, color=INK)
                c.fill = PatternFill('solid', fgColor=SOFT)
        for i in range(len(heads) + 1, ncols + 1):
            ws.cell(row=r, column=i).border = BOX
        ws.row_dimensions[r].height = 18
        r += 1
    for extra in range(16):
        for i in range(1, ncols + 1):
            ws.cell(row=r, column=i).border = BOX
        ws.row_dimensions[r].height = 18
        r += 1
    last = r - 1

    for rr in range(hrow + 1, last + 1):
        ws.cell(row=rr, column=5).number_format = 'm/d/yy'
        ws.cell(row=rr, column=6).number_format = 'm/d/yy'
        ws.cell(row=rr, column=7).value = (
            '=IF(OR(E%d="",F%d=""),"",F%d-E%d+1)' % (rr, rr, rr, rr))

    lists = wb.create_sheet('Lists')
    lists['A1'] = 'Work type'
    lists['B1'] = 'Status'
    for c in ('A1', 'B1'):
        lists[c].font = Font(bold=True, color=PURPLE)
    types = ['Research', 'Brainstorming', 'Drawing', 'CAD', 'Model',
             'Written', 'Construction', 'Testing', 'Learn']
    stat = ['Not started', 'In progress', 'Blocked', 'Complete']
    for i, v in enumerate(types, start=2):
        lists.cell(row=i, column=1, value=v)
    for i, v in enumerate(stat, start=2):
        lists.cell(row=i, column=2, value=v)
    lists.column_dimensions['A'].width = 20
    lists.column_dimensions['B'].width = 18

    dv1 = DataValidation(type='list',
                         formula1='=Lists!$A$2:$A$%d' % (len(types) + 1),
                         allow_blank=True)
    dv2 = DataValidation(type='list',
                         formula1='=Lists!$B$2:$B$%d' % (len(stat) + 1),
                         allow_blank=True)
    ws.add_data_validation(dv1)
    ws.add_data_validation(dv2)
    dv1.add('C%d:C%d' % (hrow + 1, last))
    dv2.add('D%d:D%d' % (hrow + 1, last))
    ws.freeze_panes = ws.cell(row=hrow + 1, column=len(heads) + 1)
    wb.save(path)


# ------------------------------------------------------------------ part list
def part_list(path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Part List'
    heads = ['Item', 'Part name', 'Qty', 'Material or spec',
             'Made or bought', 'Source / part number', 'Unit cost',
             'Total', 'Notes']
    widths = [6, 32, 6, 26, 15, 30, 11, 11, 34]
    _sheet_head(ws, 'Part List',
                'Every part in the design, including the fasteners. If it is '
                'not on this list it does not exist when you go to build.',
                len(heads))
    _identity_rows(ws, 5, ['Name', 'Project', 'Revision', 'Date'])
    hrow = 11
    _header_row(ws, hrow, heads, widths)
    last = hrow + 40
    _blank_rows(ws, hrow + 1, last, len(heads))
    for r in range(hrow + 1, last + 1):
        ws.cell(row=r, column=1).value = r - hrow
        ws.cell(row=r, column=1).font = Font(name='Calibri', size=9,
                                             color=INK3)
        ws.cell(row=r, column=8).value = (
            '=IF(OR(C%d="",G%d=""),"",C%d*G%d)' % (r, r, r, r))
        ws.cell(row=r, column=7).number_format = '$#,##0.00'
        ws.cell(row=r, column=8).number_format = '$#,##0.00'

    tot = last + 2
    c = ws.cell(row=tot, column=7, value='Total')
    c.font = Font(name='Calibri', size=10, bold=True, color=PURPLE)
    c.alignment = Alignment(horizontal='right')
    t = ws.cell(row=tot, column=8,
                value='=SUM(H%d:H%d)' % (hrow + 1, last))
    t.font = Font(name='Calibri', size=11, bold=True, color=INK)
    t.number_format = '$#,##0.00'
    t.border = BOX

    ws.cell(row=tot + 2, column=1,
            value='Count the fasteners. A part list that says "screws" '
                  'instead of "M3 x 12 socket head, 8 off" is not finished.'
            ).font = Font(name='Calibri', size=9, italic=True, color=INK3)

    lists = wb.create_sheet('Lists')
    lists['A1'] = 'Made or bought'
    lists['A1'].font = Font(bold=True, color=PURPLE)
    kinds = ['Make — 3D print', 'Make — laser', 'Make — CNC',
             'Make — hand/shop', 'Buy — stock', 'Buy — order',
             'Reuse — salvaged']
    for i, v in enumerate(kinds, start=2):
        lists.cell(row=i, column=1, value=v)
    lists.column_dimensions['A'].width = 22
    dv = DataValidation(type='list',
                        formula1='=Lists!$A$2:$A$%d' % (len(kinds) + 1),
                        allow_blank=True)
    ws.add_data_validation(dv)
    dv.add('E%d:E%d' % (hrow + 1, last))
    ws.freeze_panes = ws.cell(row=hrow + 1, column=1)
    wb.save(path)


# ------------------------------------------------------------ decision matrix
def decision_matrix(path):
    """Weighted criteria down the side, concepts across the top. The weight
    column is the whole point: it forces the argument about what matters
    BEFORE the argument about which idea wins."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Decision Matrix'
    NC = 4
    heads = ['Criterion', 'Weight (1\u20135)'] + \
            ['Concept %s' % c for c in 'ABCD']
    widths = [30, 12] + [14] * NC
    _sheet_head(ws, 'Decision Matrix',
                'Score each concept against each criterion from 1 to 5. The '
                'weighted totals at the bottom decide; the row you argued '
                'about longest is the one that mattered.', len(heads))
    _identity_rows(ws, 5, ['Name', 'Project', 'Date'])
    hrow = 10
    _header_row(ws, hrow, heads, widths)
    seed = ['Meets the design requirements', 'Cost', 'Time to build',
            'Safety', 'Ease of manufacture', 'Durability', 'Appearance']
    last = hrow + 12
    _blank_rows(ws, hrow + 1, last, len(heads))
    for i, c in enumerate(seed):
        ws.cell(row=hrow + 1 + i, column=1, value=c)
    for r in range(hrow + 1, last + 1):
        ws.cell(row=r, column=2).number_format = '0'
    wrow = last + 2
    ws.cell(row=wrow, column=1, value='Weighted total').font = Font(
        name='Calibri', size=10, bold=True, color=PURPLE)
    for j in range(NC):
        col = 3 + j
        L = get_column_letter(col)
        c = ws.cell(row=wrow, column=col,
                    value='=SUMPRODUCT($B$%d:$B$%d,%s%d:%s%d)'
                    % (hrow + 1, last, L, hrow + 1, L, last))
        c.font = Font(name='Calibri', size=11, bold=True, color=INK)
        c.border = BOX
        c.fill = PatternFill('solid', fgColor=SOFT)
    ws.cell(row=wrow + 2, column=1,
            value='Name the concepts in the header row. Leave the weight '
                  'blank for a criterion you decide does not apply.'
            ).font = Font(name='Calibri', size=9, italic=True, color=INK3)
    dv = DataValidation(type='whole', operator='between', formula1='1',
                        formula2='5', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add('B%d:%s%d' % (hrow + 1, get_column_letter(2 + NC), last))
    ws.freeze_panes = ws.cell(row=hrow + 1, column=2)
    wb.save(path)


# ------------------------------------------------------------------ test log
def test_log(path):
    """One instrument for every pathway that measures something: mechanical
    test data, electrical measurements, automation cycle times. One row per
    reading. The Predicted and Difference columns are what make it engineering
    rather than record-keeping."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Test Log'
    heads = ['#', 'Date', 'Test / what you measured', 'Setup and conditions',
             'Predicted', 'Measured', 'Unit', 'Difference',
             'What the difference tells you to change']
    widths = [5, 11, 30, 30, 11, 11, 8, 11, 40]
    _sheet_head(ws, 'Test and Measurement Log',
                'One row per reading. Write the prediction BEFORE you take '
                'the measurement, or the last column has nothing to say.',
                len(heads))
    _identity_rows(ws, 5, ['Name', 'Project', 'Instrument(s) used'])
    hrow = 10
    _header_row(ws, hrow, heads, widths)
    last = hrow + 40
    _blank_rows(ws, hrow + 1, last, len(heads))
    for r in range(hrow + 1, last + 1):
        ws.cell(row=r, column=1, value=r - hrow).font = Font(
            name='Calibri', size=9, color=INK3)
        ws.cell(row=r, column=2).number_format = 'm/d/yy'
        ws.cell(row=r, column=8).value = (
            '=IF(OR(E%d="",F%d=""),"",F%d-E%d)' % (r, r, r, r))
    ws.freeze_panes = ws.cell(row=hrow + 1, column=3)
    wb.save(path)


# -------------------------------------------------------------------- I/O map
def io_map(path):
    """Every input and output on an automated system, what it is wired to and
    what it means. The first document a controls engineer writes and the one
    they reach for when it stops."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'I-O Map'
    heads = ['Tag', 'Type', 'Device', 'Wired to (pin / terminal)',
             'Signal', 'Meaning when ON / high', 'Meaning when OFF / low',
             'Tested?', 'Notes']
    widths = [10, 12, 24, 20, 12, 26, 26, 9, 30]
    _sheet_head(ws, 'I/O Map',
                'Every input and output, one row each. If it is not on this '
                'sheet it is not commissioned.', len(heads))
    _identity_rows(ws, 5, ['Name', 'System', 'Controller', 'Revision'])
    hrow = 11
    _header_row(ws, hrow, heads, widths)
    last = hrow + 32
    _blank_rows(ws, hrow + 1, last, len(heads))

    lists = wb.create_sheet('Lists')
    lists['A1'] = 'Type'; lists['B1'] = 'Signal'; lists['C1'] = 'Tested'
    for c in ('A1', 'B1', 'C1'):
        lists[c].font = Font(bold=True, color=PURPLE)
    types = ['Digital input', 'Digital output', 'Analogue input',
             'Analogue output', 'Comms']
    sigs = ['24 V DC', '5 V logic', '3.3 V logic', '0\u201310 V',
            '4\u201320 mA', 'PWM', 'I2C', 'SPI', 'Serial']
    tested = ['Yes', 'No', 'Fault']
    for i, v in enumerate(types, 2): lists.cell(row=i, column=1, value=v)
    for i, v in enumerate(sigs, 2): lists.cell(row=i, column=2, value=v)
    for i, v in enumerate(tested, 2): lists.cell(row=i, column=3, value=v)
    for col, opts, letter in ((2, types, 'A'), (5, sigs, 'B'),
                              (8, tested, 'C')):
        dv = DataValidation(type='list',
                            formula1='=Lists!$%s$2:$%s$%d'
                            % (letter, letter, len(opts) + 1),
                            allow_blank=True)
        ws.add_data_validation(dv)
        L = get_column_letter(col)
        dv.add('%s%d:%s%d' % (L, hrow + 1, L, last))

    # commissioning checklist on a second tab -- same document, because you
    # commission against the map
    ck = wb.create_sheet('Commissioning')
    _sheet_head(ck, 'Commissioning Checklist',
                'Tick each line only when you have SEEN it, not when you '
                'believe it.', 3)
    _header_row(ck, 5, ['Check', 'Done', 'Notes'], [56, 8, 40])
    steps = [
        'Power off. Every wire traced against the I/O map.',
        'Grounds and commons connected; no floating references.',
        'Fuses / breakers correct for each circuit.',
        'Emergency stop tested: cuts power to every actuator.',
        'Power on with all outputs disabled. Supply voltages measured.',
        'Each INPUT forced by hand and seen to change in the controller.',
        'Each OUTPUT driven from the controller and the device seen to act.',
        'Sensor readings sanity-checked against a known value.',
        'Guards in place. Nothing moving that a hand can reach.',
        'First automatic cycle run at reduced speed, with a hand on the stop.',
        'Cycle time and error count recorded in the Test Log.',
        'Instructor sign-off before unattended running.',
    ]
    for i, st in enumerate(steps, 6):
        ck.cell(row=i, column=1, value=st).border = BOX
        ck.cell(row=i, column=2).border = BOX
        ck.cell(row=i, column=3).border = BOX
        ck.row_dimensions[i].height = 22
    dvc = DataValidation(type='list', formula1='"\u2611,\u2610"', allow_blank=True)
    ck.add_data_validation(dvc)
    dvc.add('B6:B%d' % (5 + len(steps)))
    ws.freeze_panes = ws.cell(row=hrow + 1, column=2)
    wb.save(path)


SHEETS = [
    ('BHR27-Research-Log.xlsx', research_log),
    ('BHR27-Order-Request-Form.xlsx', order_request),
    ('BHR27-Project-Gantt-Chart.xlsx', gantt),
    ('BHR27-Part-List.xlsx', part_list),
    ('BHR27-Decision-Matrix.xlsx', decision_matrix),
    ('BHR27-Test-Log.xlsx', test_log),
    ('BHR27-IO-Map-and-Commissioning.xlsx', io_map),
]


def main():
    os.makedirs(OUT, exist_ok=True)
    made = []
    for spec in DATA.DOCS:
        made.append(SD.build(_unent(spec), OUT))
    for name, fn in SHEETS:
        p = os.path.join(OUT, name)
        fn(p)
        made.append(p)
    for p in made:
        print('  %-56s %7d' % (os.path.basename(p), os.path.getsize(p)))
    print('\n%d files -> %s' % (len(made), OUT))


if __name__ == '__main__':
    main()
