# -*- coding: utf-8 -*-
"""Two outputs from one source.

1. The Classroom-shaped .xlsx, built by loading Dan's own export and writing
   ONLY the description cells, so every other cell -- the warning in A1, the
   version tag in A2, the points, the level names, the blank spacer rows --
   is byte-for-byte what Classroom produced. It goes straight back in.

2. A student-facing PDF of the same rubric, in the site's style, with the
   points shown once per level rather than repeated per row. This is the one
   that gets linked from assignment pages and the grading page.
"""

import os
import shutil
import openpyxl
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib.units import inch
from reportlab.lib import colors
from reportlab.lib.styles import ParagraphStyle
from reportlab.platypus import (BaseDocTemplate, PageTemplate, Frame,
                                Paragraph, Spacer, Table, TableStyle)
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

import rubric_data as RD
import generation as G

HERE = os.path.dirname(os.path.abspath(__file__))
import paths
SRC = {
    'project': os.path.join(paths.RUBRIC_SOURCES, 'project-rubric-classroom-export.xlsx'),
    'weekly': os.path.join(paths.RUBRIC_SOURCES, 'weekly-grade-classroom-export.xlsx'),
}
OUT = os.path.join(HERE, 'student-docs')

# ------------------------------------------------------------------- xlsx


def write_xlsx(kind, spec, out_path, sheet_title):
    wb = openpyxl.load_workbook(SRC[kind])
    ws = wb.active
    # locate each criterion block by its title cell in column A, then write
    # the five descriptions into row+4, columns B..F. Nothing else is touched.
    titles = {c.value: c.row for c in ws['A'] if c.value in
              [t for t, _, _ in spec['criteria']]}
    for title, points, descs in spec['criteria']:
        r = titles[title]
        # sanity: the points row must match Dan's export exactly
        got = [ws.cell(row=r + 2, column=c).value for c in range(2, 7)]
        assert [float(x) for x in got] == points, (title, got, points)
        for i, d in enumerate(descs):
            ws.cell(row=r + 4, column=2 + i, value=d)
    ws.title = sheet_title
    wb.save(out_path)
    return out_path


# -------------------------------------------------------------------- pdf

import paths
GF = paths.FONTS
for n, f in (('P', 'Poppins-Regular.ttf'), ('PB', 'Poppins-Bold.ttf'),
             ('PM', 'Poppins-Medium.ttf')):
    pdfmetrics.registerFont(TTFont(n, os.path.join(GF, f)))
pdfmetrics.registerFont(TTFont('L', os.path.join(GF, 'Lora-Variable.ttf')))
pdfmetrics.registerFont(TTFont('LI', os.path.join(GF,
                                                  'Lora-Italic-Variable.ttf')))

PURPLE = colors.HexColor('#6b4785')
SOFT = colors.HexColor('#f1ebf7')
INK = colors.HexColor('#262b39')
INK2 = colors.HexColor('#4f5566')
INK3 = colors.HexColor('#7d8394')
RULE = colors.HexColor('#dedbd5')
PAPER = colors.HexColor('#faf8f4')
LEVEL_TINT = [colors.HexColor('#e6f0e8'), colors.HexColor('#eef3ea'),
              colors.HexColor('#faf5e6'), colors.HexColor('#fbeee6'),
              colors.HexColor('#f8e6e6')]

PW, PH = landscape(letter)
M = 0.45 * inch
W = PW - 2 * M

st_h = ParagraphStyle('h', fontName='PB', fontSize=18, leading=21,
                      textColor=INK)
st_i = ParagraphStyle('i', fontName='L', fontSize=9.6, leading=13.4,
                      textColor=INK2, spaceBefore=3, spaceAfter=7)
st_crit = ParagraphStyle('c', fontName='PB', fontSize=9.2, leading=11.5,
                         textColor=PURPLE)
st_lvl = ParagraphStyle('l', fontName='PB', fontSize=8.2, leading=10,
                        textColor=INK)
st_pts = ParagraphStyle('p', fontName='PM', fontSize=7.4, leading=9,
                        textColor=INK3)
st_d = ParagraphStyle('d', fontName='L', fontSize=7.5, leading=9.7,
                      textColor=INK)
st_foot = ParagraphStyle('f', fontName='LI', fontSize=8.4, leading=11.5,
                         textColor=INK2, spaceBefore=8)


def _page(c, doc):
    c.saveState()
    c.setFillColor(PAPER)
    c.rect(0, 0, PW, PH, fill=1, stroke=0)
    c.setFont('P', 7.2)
    c.setFillColor(INK3)
    c.drawString(M, 0.36 * inch,
                 'Blue Hills Regional Technical School  ·  Engineering '
                 'Technology  ·  Room E-126  ·  ' + G.STAMP)
    c.drawRightString(PW - M, 0.36 * inch,
                      'The same rubric Google Classroom grades against')
    c.restoreState()


def write_pdf(spec, out_path, foot):
    doc = BaseDocTemplate(out_path, pagesize=landscape(letter),
                          leftMargin=M, rightMargin=M, topMargin=M,
                          bottomMargin=0.6 * inch, title=spec['title'],
                          author='BHR Engineering Technology')
    fr = Frame(M, 0.6 * inch, W, PH - M - 0.6 * inch, id='f',
               leftPadding=0, rightPadding=0, topPadding=0, bottomPadding=0)
    doc.addPageTemplates([PageTemplate(id='p', frames=[fr], onPage=_page)])

    F = [Paragraph(spec['title'], st_h), Paragraph(spec['intro'], st_i)]

    # header row: level name + points (points differ slightly per row in the
    # weekly grade, so show the range once and the exact figure in each cell)
    head = [Paragraph('', st_lvl)]
    for i, lv in enumerate(RD.LEVELS):
        head.append(Paragraph(lv, st_lvl))
    rows = [head]
    for title, points, descs in spec['criteria']:
        cell = [Paragraph(title.title() if title.isupper() else title,
                          st_crit)]
        for i, d in enumerate(descs):
            cell.append([Paragraph('%g pts' % points[i], st_pts),
                         Paragraph(d, st_d)])
        rows.append(cell)

    cw = [1.35 * inch] + [(W - 1.35 * inch) / 5] * 5
    t = Table(rows, colWidths=cw, repeatRows=1)
    style = [
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('BACKGROUND', (0, 0), (-1, 0), SOFT),
        ('LINEBELOW', (0, 0), (-1, 0), 1.2, PURPLE),
        ('LINEBELOW', (0, 1), (-1, -2), 0.5, RULE),
        ('BOX', (0, 0), (-1, -1), 0.7, RULE),
        ('LINEAFTER', (0, 0), (0, -1), 0.7, RULE),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4.5),
        ('LEFTPADDING', (0, 0), (-1, -1), 6),
        ('RIGHTPADDING', (0, 0), (-1, -1), 6),
    ]
    for i in range(5):
        style.append(('BACKGROUND', (1 + i, 1), (1 + i, -1), LEVEL_TINT[i]))
    t.setStyle(TableStyle(style))
    F.append(t)
    F.append(Paragraph(foot, st_foot))
    doc.build(F)
    return out_path


if __name__ == '__main__':
    os.makedirs(OUT, exist_ok=True)
    made = []
    made.append(write_xlsx('project', RD.PROJECT,
                           os.path.join(OUT, 'BHR27-Project-Rubric.xlsx'),
                           G.title('Project Rubric')))
    made.append(write_xlsx('weekly', RD.WEEKLY,
                           os.path.join(OUT, 'BHR27-Weekly-Grade-Rubric.xlsx'),
                           G.title('Weekly Grade')))
    made.append(write_pdf(
        RD.PROJECT, os.path.join(OUT, 'BHR27-Project-Rubric.pdf'),
        'Five criteria, twenty points each, one hundred in all. '
        '&ldquo;Requirements&rdquo; means the ones in the brief for this '
        'project &mdash; read it again before you hand in.'))
    made.append(write_pdf(
        RD.WEEKLY, os.path.join(OUT, 'BHR27-Weekly-Grade-Rubric.pdf'),
        'Six criteria, roughly seventeen points each. The first four are '
        'behaviour; the last two are performance. You are graded on this '
        'every shop week, and it is worth more than any single project.'))
    for m in made:
        print('%-60s %6d' % (os.path.basename(m), os.path.getsize(m)))
